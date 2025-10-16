from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.db import transaction
from datetime import timedelta
from django.db.models import Q, Sum
from django.core.mail import send_mail
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.middleware.csrf import get_token
import json
import logging
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from .models import Customer, Bike, RepairJob

logger = logging.getLogger(__name__)

from .models import (
    Customer, Bike, RepairCategory, RepairSubCategory, RepairJob,
    UserProfile, RepairItem, RepairUpdate
)
from .forms import (
    BikeForm, RepairCategoryForm, RepairSubCategoryForm, RepairJobForm,
    CustomerForm, CustomerRegisterForm, CustomerLinkForm,
    CustomerRepairJobForm, RepairDiagnosisForm, RepairItemForm,
    CustomerApprovalForm, MechanicTaskForm, CustomerWithBikeForm,
    CustomerAddBikeForm
)
from .notification_service import NotificationService
from .realtime_service import realtime_service

# Helper functions for role-based access
def is_manager(user):
    return hasattr(user, 'userprofile') and user.userprofile.role == 'manager'

def is_mechanic(user):
    return hasattr(user, 'userprofile') and user.userprofile.role in ['mechanic', 'manager']

def is_customer(user):
    return hasattr(user, 'userprofile') and user.userprofile.role == 'customer'

def manager_required(view_func):
    return user_passes_test(is_manager)(view_func)

def mechanic_required(view_func):
    return user_passes_test(is_mechanic)(view_func)

def customer_required(view_func):
    return user_passes_test(is_customer)(view_func)

@login_required
def customer_dashboard(request):
    customer = get_object_or_404(Customer, user=request.user)
    bikes = Bike.objects.filter(customer=customer)
    repairs = RepairJob.objects.filter(bike__in=bikes).select_related('bike')
    
    return render(request, 'workshop/customer_dashboard.html', {
        'customer': customer,
        'bikes': bikes,
        'repairs': repairs,
    })

def has_quality_fields():
    try:
        # רשימת השדות שמחויבים כדי שקטע בקרת האיכות יעבוד
        required_fields = [
            'quality_check_date',
            'ready_for_pickup_date',
            'available_for_pickup'
        ]

        # קבלת כל שמות השדות של המודל RepairJob
        existing_fields = [field.name for field in RepairJob._meta.get_fields()]

        # בדיקה אילו שדות חסרים
        missing_fields = [field for field in required_fields if field not in existing_fields]

        # Return True only if all fields exist
        return not missing_fields

    except Exception:
        return False



@login_required
def customer_report(request):
    try:
        customer = Customer.objects.get(user=request.user)
    except Customer.DoesNotExist:
        messages.error(request, "לא נמצא פרופיל לקוח עבור המשתמש הזה")
        return redirect('home')
    
    bikes = Bike.objects.filter(customer=customer)
    if request.method == 'POST':
        form = CustomerRepairJobForm(request.POST)
        form.fields['bike'].queryset = bikes
        if form.is_valid():
            repair = form.save(commit=False)
            if repair.bike in bikes:
                repair.save()
                form.save_m2m()
                messages.success(request, "הדיווח נשלח בהצלחה!")
                return render(request, 'workshop/customer_report_done.html')
    else:
        form = CustomerRepairJobForm()
        form.fields['bike'].queryset = bikes
    
    # הוספת קטגוריות כמו בדף המנהל
    categories = RepairCategory.objects.prefetch_related('subcategories').all()
    return render(request, 'workshop/customer_report.html', {
        'form': form,
        'categories': categories,
    })


@login_required
@user_passes_test(lambda u: is_manager(u) or is_mechanic(u))
def repair_form(request, customer_id=None):
    selected_customer = None
    if customer_id:
        try:
            selected_customer = Customer.objects.get(id=customer_id)
        except Customer.DoesNotExist:
            messages.error(request, "לקוח לא נמצא")
            return redirect('customer_list')
    
    if request.method == 'POST':
        form = RepairJobForm(request.POST)
        if form.is_valid():
            try:
                # שמירת התיקון
                repair = form.save(commit=False)
                repair.status = 'reported'  # סטטוס ראשוני
                repair.save()
                form.save_m2m()  # שמירת קטגוריות משנה
                
                # אם המנהל הזין אבחון, נעביר אותו ישירות לדף האבחון המפורט
                if is_manager(request.user) and repair.diagnosis:
                    messages.success(request, "התיקון נוצר בהצלחה! כעת הוסף פעולות תיקון ספציפיות עם מחירים.")
                    return redirect('repair_diagnosis', repair_id=repair.id)
                else:
                    messages.success(request, "התיקון נוצר בהצלחה!")
                    if is_manager(request.user):
                        return redirect('manager_dashboard')
                    else:
                        return redirect('home')
                        
            except Exception as e:
                messages.error(request, f"שגיאה ביצירת התיקון: {str(e)}")
    else:
        form = RepairJobForm()
    categories = RepairCategory.objects.prefetch_related('subcategories').all()
    return render(request, 'workshop/repair_form.html', {
        'form': form,
        'categories': categories,
        'selected_customer': selected_customer,
    })


@login_required
@user_passes_test(lambda u: is_manager(u) or is_mechanic(u))
def bike_form(request):
    if request.method == 'POST':
        form = BikeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "האופניים נוספו בהצלחה!")
            return redirect('home')
    else:
        form = BikeForm()
    return render(request, 'workshop/bike_form.html', {'form': form})


@login_required
@user_passes_test(lambda u: is_manager(u) or is_mechanic(u))
def customer_form(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "הלקוח נוצר בהצלחה!")
            return redirect('home')
    else:
        form = CustomerForm()
    return render(request, 'workshop/customer_form.html', {'form': form})



def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Ensure we have the required fields
        if not username or not password:
            messages.error(request, 'נא למלא שם משתמש וסיסמה')
            return render(request, 'workshop/login.html')
            
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            # Get the next URL if available
            next_url = request.GET.get('next', '/')
            return redirect(next_url)
        else:
            messages.error(request, 'שם משתמש או סיסמה שגויים')
            return render(request, 'workshop/login.html')
    return render(request, 'workshop/login.html')


def home(request):
    context = {}
    
    try:
        if request.user.is_authenticated and hasattr(request.user, 'userprofile'):
            role = request.user.userprofile.role
            
            if role == 'customer':
                # עבור לקוח - הצג תיקונים שלו
                try:
                    customer = request.user.customer
                    all_repairs = RepairJob.objects.filter(
                        bike__customer=customer
                    ).select_related('bike').order_by('-created_at')
                    
                    # תיקונים הממתינים לאישור
                    pending_approval = all_repairs.filter(status='diagnosed')[:5]
                    
                    # תיקונים אחרונים
                    recent_repairs = all_repairs[:5]
                    
                    # ADD THESE NEW LINES FOR CUSTOMER HOME PAGE:
                    # תיקונים פעילים (לכרטיס "התיקונים שלי")
                    customer_active_repairs = all_repairs.filter(
                        status__in=['pending', 'in_progress', 'approved_partial', 'approved_full', 'diagnosed']
                    )[:5]
                    
                    # סטטיסטיקות שנתיות
                    current_year = timezone.now().year
                    customer_yearly_repairs_count = all_repairs.filter(
                        created_at__year=current_year
                    ).count()
                    
                    # סה"כ הוצאות (מתיקונים עם מחיר)
                    customer_total_spent = 0
                    for repair in all_repairs:
                        if hasattr(repair, 'repair_items'):
                            customer_total_spent += sum(
                                item.price for item in repair.repair_items.all() 
                                if item.is_approved_by_customer
                            )
                    
                    # זמן ממוצע לתיקון
                    completed_repairs = all_repairs.filter(status__in=['completed', 'delivered'])
                    if completed_repairs.exists():
                        total_days = 0
                        count = 0
                        for repair in completed_repairs:
                            if hasattr(repair, 'completed_at') and repair.completed_at:
                                days = (repair.completed_at - repair.created_at).days
                                total_days += days
                                count += 1
                            elif hasattr(repair, 'diagnosed_at') and repair.diagnosed_at:
                                # אם אין completed_at, נשתמש בזמן אבחון כהערכה
                                days = (repair.diagnosed_at - repair.created_at).days + 3  # הערכה
                                total_days += days
                                count += 1
                        customer_avg_days = total_days // count if count > 0 else 0
                    else:
                        customer_avg_days = 0
                    
                    # מספר אופניים של הלקוח
                    customer_bikes_count = customer.bikes.count() if hasattr(customer, 'bikes') else 0
                    
                    context.update({
                        'recent_repairs': recent_repairs,
                        'pending_approval': pending_approval,
                        'customer': customer,
                        # NEW DATA FOR CUSTOMER HOME PAGE:
                        'customer_active_repairs': customer_active_repairs,
                        'customer_yearly_repairs_count': customer_yearly_repairs_count,
                        'customer_total_spent': customer_total_spent,
                        'customer_avg_days': customer_avg_days,
                        'customer_bikes_count': customer_bikes_count,
                    })
                except Customer.DoesNotExist:
                    context['no_customer_profile'] = True
                
                return render(request, 'workshop/customer_home.html', context)
                    
            elif role == 'mechanic':
                # عبور مכונאי - هצג تیקונים המוقצים אליו
                assigned_repairs = RepairJob.objects.filter(
                    assigned_mechanic=request.user,
                    status='in_progress'
                ).select_related('bike', 'bike__customer')[:5]
                
                context.update({
                    'assigned_repairs': assigned_repairs,
                })
                
                return render(request, 'workshop/mechanic_home.html', context)
                
            elif role == 'manager':
                # عبור מנהל - تقציר مהیر + تیקונים موقצים (כמו مכונאי)
                pending_diagnosis = RepairJob.objects.filter(status='reported').count()
                pending_approval = RepairJob.objects.filter(status='diagnosed').count()
                in_progress = RepairJob.objects.filter(status='in_progress').count()
                blocked_tasks_count = RepairJob.objects.filter(is_stuck=True).count()
                ready_for_pickup = RepairJob.objects.filter(status='quality_approved').count()

                # הכנסה צפויה מתיקונים ממתינים לאישור
                expected_revenue = RepairJob.objects.filter(
                    status__in=['diagnosed', 'approved_partial', 'approved_full']
                ).aggregate(
                    total=Sum('repair_items__price')
                )['total'] or 0
                
                # חישוב יעילות מוסך (אחוז תיקונים שהושלמו השבוע)
                week_ago = timezone.now() - timedelta(days=7)
                
                total_repairs_week = RepairJob.objects.filter(created_at__gte=week_ago).count()
                completed_week = RepairJob.objects.filter(
                    created_at__gte=week_ago,
                    status__in=['quality_approved', 'delivered', 'completed']
                ).count()
                
                efficiency = int((completed_week / total_repairs_week * 100)) if total_repairs_week > 0 else 0
                
                # תיקונים מוקצים למנהל כמכונאי
                assigned_repairs = RepairJob.objects.filter(
                    assigned_mechanic=request.user,
                    status='in_progress'
                ).select_related('bike', 'bike__customer')[:5]
                
                context.update({
                    'pending_diagnosis_count': pending_diagnosis,
                    'pending_approval_count': pending_approval,
                    'in_progress_count': in_progress,
                    'assigned_repairs': assigned_repairs,
                    'blocked_tasks_count': blocked_tasks_count,
                    'ready_for_pickup_count': ready_for_pickup,
                    'expected_revenue': expected_revenue,
                    'efficiency': efficiency,
                })
                
                return render(request, 'workshop/manager_home_react.html', context)
    except Exception:
        context['error'] = "בעיה בטעינת הנתונים"
    
    # Default fallback to original home template for guest users
    return render(request, 'workshop/home.html', context)

def user_logout(request):
    logout(request)
    return redirect("login")

@login_required
@manager_required
def category_list(request):
    categories = RepairCategory.objects.prefetch_related('subcategories').all()
    return render(request, 'workshop/category_list.html', {'categories': categories})


@login_required
@manager_required
def category_list_react(request):
    """React-based category management interface"""
    return render(request, 'workshop/category_list_react.html')


@login_required
@manager_required
def category_create(request):
    if request.method == 'POST':
        form = RepairCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "הקטגוריה נוצרה בהצלחה!")
            return redirect('category_list')
    else:
        form = RepairCategoryForm()
    return render(request, 'workshop/category_form.html', {'form': form})

@login_required
@manager_required
def subcategory_create(request):
    if request.method == 'POST':
        form = RepairSubCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "תת-הקטגוריה נוצרה בהצלחה!")
            return redirect('category_list')
    else:
        form = RepairSubCategoryForm()
    return render(request, 'workshop/subcategory_form.html', {'form': form})


@login_required
@manager_required
def subcategory_create_react(request):
    """React-based subcategory management interface"""
    return render(request, 'workshop/subcategory_form_react.html')


def register(request):
    if request.method == 'POST':
        form = CustomerRegisterForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                login(request, user)
                
                # בדיקה אם חובר לקוח קיים
                if hasattr(form, 'existing_customer') and form.cleaned_data.get('link_existing'):
                    messages.success(request, f"נרשמת בהצלחה! חשבונך חובר ללקוח הקיים במערכת. כעת תוכל לראות את כל האופניים והתיקונים שלך.")
                else:
                    messages.success(request, "נרשמת בהצלחה! ברוך הבא למערכת המוסך.")
                
                return redirect('home')
            except Exception as e:
                messages.error(request, f"שגיאה ברישום: {str(e)}")
                # Registration error occurred
        else:
            messages.error(request, "יש שגיאות בטופס. אנא תקן ונסה שוב.")
    else:
        form = CustomerRegisterForm()
    return render(request, 'workshop/register.html', {'form': form})

@login_required
@manager_required
def customer_list(request):
    """רשימת כל הלקוחות - עם חיפוש ו-pagination"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # קבלת פרמטרי חיפוש וסינון
    search_query = request.GET.get('search', '').strip()
    filter_type = request.GET.get('filter', 'all')  # all, with_user, without_user
    
    # בניית שאילתה בסיסית
    customers = Customer.objects.select_related('user').prefetch_related('bikes')
    
    # סינון לפי סוג לקוח
    if filter_type == 'with_user':
        customers = customers.filter(user__isnull=False)
    elif filter_type == 'without_user':
        customers = customers.filter(user__isnull=True)
    
    # חיפוש טקסט
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )
    
    # מיון לפי שם
    customers = customers.order_by('name')
    
    # Pagination - 25 לקוחות לעמוד
    paginator = Paginator(customers, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # סטטיסטיקות
    total_customers = Customer.objects.count()
    customers_with_user_qs = Customer.objects.filter(user__isnull=False)
    customers_without_user_qs = Customer.objects.filter(user__isnull=True)
    customers_with_user = customers_with_user_qs.count()
    customers_without_user = customers_without_user_qs.count()
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'filter_type': filter_type,
        'total_customers': total_customers,
        'customers_with_user': customers_with_user_qs,  # queryset for iteration
        'customers_without_user': customers_without_user_qs,  # queryset for iteration
        'customers_with_user_count': customers_with_user,  # count for display
        'customers_without_user_count': customers_without_user,  # count for display
        'filtered_count': paginator.count,
    }
    return render(request, 'workshop/customer_list.html', context)

@login_required
@manager_required
def link_customer_to_user(request):
    """מנהל יכול לחבר לקוח קיים למשתמש קיים"""
    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        user_id = request.POST.get('user_id')
        
        try:
            customer = Customer.objects.get(id=customer_id, user__isnull=True)
            user = User.objects.get(id=user_id)
            
            # בדיקה שהמשתמש עדיין לא מחובר ללקוח אחר
            if hasattr(user, 'customer'):
                messages.error(request, f"המשתמש {user.username} כבר מחובר ללקוח אחר")
            else:
                customer.user = user
                customer.save()
                messages.success(request, f"הלקוח {customer.name} חובר בהצלחה למשתמש {user.username}")
                
        except (Customer.DoesNotExist, User.DoesNotExist):
            messages.error(request, "שגיאה בחיבור הלקוח למשתמש")
    
    # הצגת לקוחות ללא משתמש ומשתמשים ללא לקוח
    customers_without_user = Customer.objects.filter(user__isnull=True)
    users_without_customer = User.objects.filter(customer__isnull=True, userprofile__role='customer')
    
    context = {
        'customers_without_user': customers_without_user,
        'users_without_customer': users_without_customer,
    }
    return render(request, 'workshop/link_customer_user.html', context)

@login_required
@manager_required
def manager_dashboard(request):
    """דשבורד למנהל - מציג תיקונים שדורשים אבחון או עדכון"""
    
    

    try:
        # תיקונים שהמכונאי סימן כתקועים
        stuck_repairs = RepairJob.objects.filter(
            is_stuck=True
        ).select_related('bike', 'bike__customer', 'assigned_mechanic').prefetch_related('repair_items')
        
        # תיקונים לפי סטטוס
        pending_diagnosis = RepairJob.objects.filter(status='reported').select_related('bike', 'bike__customer')
        pending_approval = RepairJob.objects.filter(status='diagnosed').select_related('bike', 'bike__customer')
        partially_approved = RepairJob.objects.filter(status='partially_approved').select_related('bike', 'bike__customer')
        
        
        # הפרדה נכונה: תיקונים מאושרים ממתינים להקצאת מכונאי vs בביצוע
        approved_waiting_for_mechanic = RepairJob.objects.filter(
            status__in=['approved', 'partially_approved'], 
            assigned_mechanic__isnull=True
        ).select_related('bike', 'bike__customer').prefetch_related('repair_items')
        
        in_progress = RepairJob.objects.filter(
            status='in_progress'
        ).select_related('bike', 'bike__customer', 'assigned_mechanic').prefetch_related('repair_items')
        
        # תיקונים מאושרים עם מכונאי מוקצה אבל עדיין לא התחילו
        approved_with_mechanic = RepairJob.objects.filter(
            status__in=['approved', 'partially_approved'],
            assigned_mechanic__isnull=False
        ).select_related('bike', 'bike__customer', 'assigned_mechanic').prefetch_related('repair_items')
        
        # תיקונים הממתינים לבדיקת איכות - רק אם השדות קיימים
        awaiting_quality_check = []
        
        has_quality = has_quality_fields()
        
        
        
        if has_quality:
            awaiting_quality_check = RepairJob.objects.filter(status='awaiting_quality_check').select_related('bike', 'bike__customer', 'assigned_mechanic')
        
        # קטגוריות חדשות - ויזואליות בלבד כרגע
        repairs_work_completed = []  # אופניים שנאספו על ידי הלקוח
        repairs_not_collected = []   # אופניים שטרם נאספו
        
        if has_quality:
            # אופניים שנאספו על ידי הלקוח (סטטוס completed/delivered)
            repairs_work_completed = RepairJob.objects.filter(
                status__in=['completed', 'delivered']
            ).select_related('bike', 'bike__customer')
            
            # אופניים שטרם נאספו (quality_approved - מוכנים לאיסוף)
            repairs_not_collected = RepairJob.objects.filter(
                status='quality_approved'
            ).select_related('bike', 'bike__customer')
            
        
        # ספירה מתוקנת
        waiting_to_start_count = 0
        actively_working_count = 0
        
        # ספירה של תיקונים מאושרים עם מכונאי
        for repair in approved_with_mechanic:
            if repair.progress_percentage == 0 and not repair.is_effectively_stuck:
                waiting_to_start_count += 1
            elif repair.progress_percentage > 0 and repair.progress_percentage < 100 and not repair.is_effectively_stuck:
                actively_working_count += 1
        
        # ספירה של תיקונים בביצוע
        for repair in in_progress:
            if repair.progress_percentage == 0 and not repair.is_effectively_stuck:
                waiting_to_start_count += 1
            elif repair.progress_percentage > 0 and repair.progress_percentage < 100 and not repair.is_effectively_stuck:
                actively_working_count += 1
        
        blocked_tasks_count = stuck_repairs.count()
        
        context = {
            'stuck_repairs': stuck_repairs,
            'pending_diagnosis': pending_diagnosis,
            'pending_approval': pending_approval,
            'partially_approved': partially_approved,
            'approved_waiting_for_mechanic': approved_waiting_for_mechanic,
            'approved_with_mechanic': approved_with_mechanic,
            'in_progress': in_progress,
            'awaiting_quality_check': awaiting_quality_check,
            'repairs_work_completed': repairs_work_completed,
            'repairs_not_collected': repairs_not_collected,
            'waiting_to_start_count': waiting_to_start_count,
            'actively_working_count': actively_working_count,
            'blocked_tasks_count': blocked_tasks_count,
        }
        
        
        return render(request, 'workshop/manager_dashboard.html', context)
    
    except Exception as e:
        # לוג השגיאה ותצוגת דף ריק
        # Manager dashboard error occurred
        
        context = {
            'error': str(e),
            'stuck_repairs': [],
            'pending_diagnosis': [],
            'pending_approval': [],
            'partially_approved': [],
            'in_progress': [],
            'awaiting_quality_check': [],
            'repairs_work_completed': [],
            'repairs_not_collected': [],
            'waiting_to_start_count': 0,
            'actively_working_count': 0,
            'blocked_tasks_count': 0,
            'all_repairs_count': 0,
        }
        
        
        return render(request, 'workshop/manager_dashboard.html', context)


@login_required
@manager_required
def manager_dashboard_react(request):
    """Manager dashboard with Vite-built React component (Phase 1 migration)"""
    return render(request, 'workshop/manager_dashboard.html')


@login_required
@manager_required
def manager_repair_detail(request, repair_id):
    """Manager repair detail page for responding to stuck repairs"""
    try:
        repair = get_object_or_404(RepairJob, id=repair_id)
        
        context = {
            'repair': repair,
            'is_manager': True,
        }
        
        return render(request, 'workshop/manager_repair_detail.html', context)
        
    except Exception as e:
        messages.error(request, f"שגיאה בטעינת תיקון: {str(e)}")
        return redirect('manager_dashboard')


@login_required
@manager_required
def repair_diagnosis(request, repair_id):
    """מנהל מוסיף אבחון ופרטי תיקון"""
    repair_job = get_object_or_404(RepairJob, id=repair_id)
    
    # מאפשר עריכה של אבחון עד לשלב אישור הלקוח
    if repair_job.status not in ['reported', 'diagnosed']:
        messages.error(request, 'לא ניתן לערוך אבחון זה יותר')
        return redirect('manager_dashboard')
    
    # משתנה לבדוק אם זה אבחון חדש או עריכה
    is_editing = repair_job.status == 'diagnosed'
    
    if request.method == 'POST':
        diagnosis_form = RepairDiagnosisForm(request.POST, instance=repair_job)
        
        # טיפול בפעולות התיקון
        repair_items_data = []
        i = 0
        while f'item_{i}_description' in request.POST:
            description = request.POST.get(f'item_{i}_description', '').strip()
            price = request.POST.get(f'item_{i}_price', '').strip()
            
            if description and price:
                try:
                    price = float(price)
                    repair_items_data.append({'description': description, 'price': price})
                except ValueError:
                    messages.error(request, f'מחיר לא תקין עבור פעולה: {description}')
                    return render(request, 'workshop/repair_diagnosis.html', {
                        'repair_job': repair_job,
                        'diagnosis_form': diagnosis_form,
                    })
            i += 1
        
        if diagnosis_form.is_valid() and repair_items_data:
            with transaction.atomic():
                # שמירת האבחון
                repair_job = diagnosis_form.save(commit=False)
                repair_job.status = 'diagnosed'
                repair_job.diagnosed_at = timezone.now()
                repair_job.save()
                
                # יצירת פעולות התיקון
                for item_data in repair_items_data:
                    RepairItem.objects.create(
                        repair_job=repair_job,
                        description=item_data['description'],
                        price=item_data['price']
                    )
                
                # הוספת עדכון
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message=f"נוסף אבחון וכתב כמויות. סה\"ג פעולות: {len(repair_items_data)}",
                    is_visible_to_customer=True
                )
                
                # בדיקה אם לשלוח התראה ללקוח
                send_notification = request.POST.get('send_notification') == 'on'
                notification_sent = False
                
                if send_notification:
                    try:
                        # שליחת התראה ללקוח באמצעות מערכת ההתראות החדשה
                        NotificationService.notify_approval_needed(repair_job)
                        notification_sent = True
                        
                        # שליחת התראת push נוספת עבור אישור נדרש
                        from .push_service import push_service
                        items_count = repair_job.repair_items.count()
                        push_service.send_approval_needed_notification(
                            repair_job.bike.customer, 
                            repair_job, 
                            items_count
                        )
                        
                        # שליחת התראה גם במערכת הישנה לתאימות לאחור
                        total_price = sum(item_data['price'] for item_data in repair_items_data)
                        send_customer_notification(
                            repair_job, 
                            'diagnosis_ready', 
                            f"סה\"ג מחיר משוער: ₪{total_price:.2f}",
                            user=request.user
                        )
                    except Exception as e:
                        messages.warning(request, f'אבחון נשמר אך שליחת ההתראה נכשלה: {str(e)}')
                
                # הודעות הצלחה
                if is_editing:
                    if notification_sent:
                        messages.success(request, 'אבחון עודכן בהצלחה! הלקוח קיבל התראה על השינויים.')
                    else:
                        messages.success(request, 'אבחון עודכן בהצלחה (ללא שליחת התראה).')
                else:
                    if notification_sent:
                        messages.success(request, 'אבחון נשמר בהצלחה והלקוח קיבל התראה. כעת הלקוח יכול לאשר את הפעולות.')
                    else:
                        messages.success(request, 'אבחון נשמר בהצלחה (ללא שליחת התראה).')
                        
                return redirect('manager_dashboard')
        else:
            if not repair_items_data:
                messages.error(request, 'יש להוסיף לפחות פעולת תיקון אחת')
    else:
        diagnosis_form = RepairDiagnosisForm(instance=repair_job)
    
    return render(request, 'workshop/repair_diagnosis.html', {
        'repair_job': repair_job,
        'diagnosis_form': diagnosis_form,
        'is_editing': is_editing,
    })

@login_required
def customer_approval(request, repair_id):
    """לקוח מאשר פעולות תיקון ספציפיות"""
    repair_job = get_object_or_404(RepairJob, id=repair_id)
    
    # בדיקת הרשאות - רק הלקוח או מנהל יכולים לגשת
    if not (is_manager(request.user) or 
            (hasattr(repair_job.bike.customer, 'user') and 
             repair_job.bike.customer.user == request.user)):
        raise PermissionDenied("אין לך הרשאה לצפות בתיקון זה")
    
    if repair_job.status not in ['diagnosed', 'partially_approved']:
        messages.error(request, 'תיקון זה לא זמין לאישור כרגע')
        return redirect('home')
    
    if request.method == 'POST':
        approval_form = CustomerApprovalForm(repair_job=repair_job, data=request.POST)
        
        if approval_form.is_valid():
            approved_items = approval_form.cleaned_data['approved_items']
            
            with transaction.atomic():
                # עדכון סטטוס הפעולות
                for item in repair_job.repair_items.all():
                    item.is_approved_by_customer = item in approved_items
                    item.save()
                
                # עדכון סטטוס התיקון
                approved_count = len(approved_items)
                total_count = repair_job.repair_items.count()
                
                if approved_count == 0:
                    repair_job.status = 'diagnosed'  # לא אושר כלום
                elif approved_count == total_count:
                    repair_job.status = 'approved'  # אושר הכל
                    repair_job.approved_at = timezone.now()
                    
                    # Send real-time notification to managers about approved repair
                    realtime_service.send_to_group(
                        "managers",
                        "repair_approved",
                        {
                            'repair_id': repair_job.id,
                            'customer_name': repair_job.bike.customer.name,
                            'bike_info': f"{repair_job.bike.brand} {repair_job.bike.model}",
                            'approved_count': approved_count,
                            'total_count': total_count,
                            'message': f"לקוח אישר תיקון: {repair_job.bike.brand} {repair_job.bike.model}"
                        }
                    )
                    
                else:
                    repair_job.status = 'partially_approved'  # אישור חלקי
                    
                    # Send real-time notification about partial approval
                    realtime_service.send_to_group(
                        "managers",
                        "repair_partially_approved",
                        {
                            'repair_id': repair_job.id,
                            'customer_name': repair_job.bike.customer.name,
                            'bike_info': f"{repair_job.bike.brand} {repair_job.bike.model}",
                            'approved_count': approved_count,
                            'total_count': total_count,
                            'message': f"אישור חלקי: {approved_count}/{total_count} פעולות"
                        }
                    )
                
                repair_job.save()
                
                # הוספת עדכון
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message=f"לקוח אישר {approved_count} מתוך {total_count} פעולות",
                    is_visible_to_customer=True
                )
                
                if approved_count > 0:
                    messages.success(request, f'אושרו {approved_count} פעולות. התיקון יועבר לביצוע.')
                else:
                    messages.info(request, 'לא אושרו פעולות. ניתן לעדכן את האישור בעתיד.')
                
                return redirect('home')
        else:
            # הוספת הודעת שגיאה למשתמש
            messages.error(request, 'היתה בעיה בטופס. אנא בדוק את הבחירות שלך.')
    else:
        approval_form = CustomerApprovalForm(repair_job=repair_job)
    
    return render(request, 'workshop/customer_approval_react.html', {
        'repair_job': repair_job,
        'approval_form': approval_form,
    })

@login_required
@manager_required
def assign_mechanic(request, repair_id):
    """מנהל מקצה מכונאי לתיקון שאושר"""
    repair_job = get_object_or_404(RepairJob, id=repair_id)
    
    if repair_job.status not in ['approved', 'partially_approved']:
        messages.error(request, 'תיקון זה לא זמין להקצאה כרגע')
        return redirect('manager_dashboard')
    
    if request.method == 'POST':
        mechanic_id = request.POST.get('mechanic_id')
        if mechanic_id:
            try:
                if mechanic_id == 'auto_assign':
                    # הקצאה אוטומטית - מצא את המכונאי עם הכי מעט עבודות פעילות
                    # אבל לא כולל מנהלים (המנהל יבחר את עצמו אם הוא רוצה)
                    mechanics = User.objects.filter(userprofile__role='mechanic')
                    
                    # חישוב עומס עבודה לכל מכונאי
                    best_mechanic = None
                    min_workload = float('inf')
                    
                    for mechanic in mechanics:
                        # ספירת תיקונים פעילים
                        active_repairs = RepairJob.objects.filter(
                            assigned_mechanic=mechanic,
                            status='in_progress'
                        ).count()
                        
                        # ספירת תיקונים תקועים (עומס נוסף)
                        stuck_repairs = RepairJob.objects.filter(
                            assigned_mechanic=mechanic,
                            is_stuck=True
                        ).count()
                        
                        # עומס כולל (תיקונים רגילים + תיקונים תקועים מכפילים)
                        total_workload = active_repairs + (stuck_repairs * 2)
                        
                        if total_workload < min_workload:
                            min_workload = total_workload
                            best_mechanic = mechanic
                    
                    if best_mechanic:
                        mechanic = best_mechanic
                        auto_assigned = True
                    else:
                        messages.error(request, 'לא נמצא מכונאי זמין להקצאה אוטומטית')
                        return redirect('assign_mechanic', repair_id=repair_id)
                else:
                    # הקצאה ספציפית למכונאי
                    mechanic = User.objects.get(id=mechanic_id, userprofile__role__in=['mechanic', 'manager'])
                    auto_assigned = False
                
                repair_job.assigned_mechanic = mechanic
                repair_job.status = 'in_progress'
                repair_job.save()
                
                # Send real-time notification to assigned mechanic
                realtime_service.mechanic_assigned(repair_job, mechanic)
                
                # הודעת עדכון מותאמת לפי סוג ההקצאה
                if auto_assigned:
                    message = f"התיקון הוקצה אוטומטית למכונאי הזמין: {mechanic.get_full_name() or mechanic.username}"
                    success_message = f'התיקון הוקצה אוטומטית למכונאי הזמין: {mechanic.get_full_name() or mechanic.username}'
                else:
                    message = f"התיקון הוקצה למכונאי: {mechanic.get_full_name() or mechanic.username}"
                    success_message = f'התיקון הוקצה למכונאי {mechanic.get_full_name() or mechanic.username}'
                
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message=message,
                    is_visible_to_customer=True
                )
                
                messages.success(request, success_message)
                return redirect('manager_dashboard')
            except User.DoesNotExist:
                messages.error(request, 'מכונאי לא נמצא')
    
    # רשימת מכונאים זמינים (כולל מנהלים) עם מידע על עומס עבודה
    mechanics = User.objects.filter(userprofile__role__in=['mechanic', 'manager'])
    
    # הוספת מידע על עומס עבודה לכל מכונאי
    for mechanic in mechanics:
        mechanic.active_repairs_count = RepairJob.objects.filter(
            assigned_mechanic=mechanic,
            status='in_progress'
        ).count()
        
        mechanic.stuck_repairs_count = RepairJob.objects.filter(
            assigned_mechanic=mechanic,
            is_stuck=True
        ).count()
        
        mechanic.total_workload = mechanic.active_repairs_count + (mechanic.stuck_repairs_count * 2)
    
    # מיון לפי עומס עבודה (הכי מעט עבודה ראשון)
    mechanics = sorted(mechanics, key=lambda m: m.total_workload)
    
    # הוספת פעולות מאושרות (לא צריך - כבר יש property)
    # repair_job.approved_items כבר קיים כ-property במודל
    
    return render(request, 'workshop/assign_mechanic.html', {
        'repair_job': repair_job,
        'mechanics': mechanics,
    })

@login_required
@mechanic_required  
def mechanic_dashboard(request):
    """דשבורד למכונאי - מציג תיקונים שהוקצו אליו"""
    try:
        assigned_repairs = RepairJob.objects.filter(
            assigned_mechanic=request.user,
            status='in_progress'
        ).select_related('bike', 'bike__customer').prefetch_related('repair_items')
        
        # הוספת מידע על התקדמות לכל תיקון (לא צריך - יש properties במודל RepairJob)
        for repair in assigned_repairs:
            # כל הפעולות הבאות כבר קיימות כ-properties במודל RepairJob
            repair.pending_items = repair.repair_items.filter(is_approved_by_customer=True, status='pending')
            
            repair.approved_count = repair.approved_items.count()
            repair.completed_count = repair.completed_items.count()
            repair.blocked_count = repair.blocked_items.count()
            repair.pending_count = repair.pending_items.count()
            # progress_percentage הוא property - לא צריך להגדיר אותו ידנית
            
            # הוספת שדות stuck באופן בטוח
            repair.is_stuck = getattr(repair, 'is_stuck', False)
            repair.stuck_reason = getattr(repair, 'stuck_reason', '')
        
        return render(request, 'workshop/mechanic_dashboard.html', {
            'assigned_repairs': assigned_repairs,
        })
    except Exception as e:
        from django.http import HttpResponse
        return HttpResponse(f"שגיאה: {str(e)}")  


@login_required
@mechanic_required
def mechanic_dashboard_react(request):
    """React-based mechanic dashboard"""
    return render(request, 'workshop/mechanic_dashboard_react.html')


@login_required
@mechanic_required
def mechanic_task_completion(request, repair_id):
    """מכונאי מסמן פעולות כמושלמות"""
    repair_job = get_object_or_404(RepairJob, id=repair_id, assigned_mechanic=request.user)
    
    if repair_job.status != 'in_progress':
        messages.error(request, 'תיקון זה לא זמין לעדכון כרגע')
        return redirect('mechanic_dashboard')
    
    if request.method == 'POST':
        # לא משתמשים בטופס אלא קוראים ישירות מ-request.POST
        general_notes = request.POST.get('general_notes', '')
        
        with transaction.atomic():
            # עדכון סטטוס והערות לכל פעולה
            approved_items = repair_job.repair_items.filter(is_approved_by_customer=True)
            
            for item in approved_items:
                # קריאת סטטוס הפעולה
                status_field = f'item_status_{item.id}'
                if status_field in request.POST:
                    item_status = request.POST[status_field]
                    
                    # Check if status changed to send real-time updates
                    old_status = item.status
                    
                    if item_status == 'completed':
                        item.status = 'completed'
                        item.is_completed = True
                        item.completed_by = request.user
                        item.completed_at = timezone.now()
                        
                        # Send real-time progress update if newly completed
                        if old_status != 'completed':
                            realtime_service.repair_item_completed(item, request.user)
                            
                    elif item_status == 'blocked':
                        item.status = 'blocked'
                        item.is_completed = False
                    else:  # pending
                        item.status = 'pending'
                        item.is_completed = False
                
                # עדכון הערות
                notes_field = f'notes_{item.id}'
                if notes_field in request.POST:
                    item_notes = request.POST[notes_field].strip()
                    if item_notes:
                        item.notes = item_notes
                
                item.save()
            
            # ספירת פעולות מושלמות
            completed_count = approved_items.filter(status='completed').count()
            
            # בדיקה אם כל הפעולות שאושרו הושלמו
            if approved_items.count() == completed_count and completed_count > 0:
                repair_job.status = 'awaiting_quality_check'
                repair_job.save()
                
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message="התיקון הושלם והועבר לבדיקת איכות על ידי המנהל.",
                    is_visible_to_customer=True
                )
                
                # Send real-time notification to managers
                realtime_service.quality_check_ready(repair_job)
                
                messages.success(request, 'התיקון הושלם והועבר לבדיקת איכות!')
            else:
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message=f"עודכנו סטטוס והערות של פעולות התיקון",
                    is_visible_to_customer=True
                )
                
                messages.success(request, f'העדכונים נשמרו בהצלחה')
            
            return redirect('mechanic_dashboard')
    else:
        task_form = MechanicTaskForm(repair_job=repair_job)
    
    # הוספת מידע נוסף לתיקון (לא צריך - יש properties במודל)
    # repair_job.approved_items, completed_items, progress_percentage כבר קיימים כ-properties
    repair_job.pending_items = repair_job.repair_items.filter(is_completed=True)
    repair_job.approved_count = repair_job.approved_items.count()
    repair_job.completed_count = repair_job.completed_items.count()
    
    return render(request, 'workshop/mechanic_task_completion.html', {
        'repair_job': repair_job,
        'task_form': task_form,
    })

@login_required
def repair_status(request, repair_id):
    """צפייה בסטטוס תיקון - זמין לכל המשתמשים המורשים"""
    repair_job = get_object_or_404(RepairJob, id=repair_id)
    
    # בדיקת הרשאות
    if not (is_manager(request.user) or 
            is_mechanic(request.user) or
            (hasattr(repair_job.bike.customer, 'user') and 
             repair_job.bike.customer.user == request.user)):
        raise PermissionDenied("אין לך הרשאה לצפות בתיקון זה")
    
    # טיפול בתגובת מנהל לתיקון תקוע
    if request.method == 'POST' and is_manager(request.user):
        action = request.POST.get('action')
        if action == 'manager_response':
            response = request.POST.get('manager_response', '').strip()
            mark_resolved = request.POST.get('mark_resolved') == 'true'
            
            if response:
                # עדכון התגובה
                repair_job.manager_response = response
                
                if mark_resolved:
                    # סימון כנפתר
                    if repair_job.is_stuck:
                        repair_job.stuck_resolved = True
                        repair_job.is_stuck = False
                    
                    # פתיחת פעולות תקועות (החזרה לסטטוס ממתין)
                    blocked_items = repair_job.repair_items.filter(is_approved_by_customer=True, status='blocked')
                    for item in blocked_items:
                        item.status = 'pending'
                        item.notes = item.notes + f"\n[מנהל פתר: {response}]" if item.notes else f"[מנהל פתר: {response}]"
                        item.save()
                    
                    # יצירת עדכון למעקב
                    RepairUpdate.objects.create(
                        repair_job=repair_job,
                        user=request.user,
                        message=f"מנהל פתר את התקיעות: {response}",
                        is_visible_to_customer=False
                    )
                    
                    messages.success(request, 'התקיעות נפתרה בהצלחה! המכונאי יוכל להמשיך עבודה.')
                else:
                    # יצירת עדכון למעקב
                    RepairUpdate.objects.create(
                        repair_job=repair_job,
                        user=request.user,
                        message=f"מנהל השיב על התקיעות: {response}",
                        is_visible_to_customer=False
                    )
                    
                    messages.success(request, 'התגובה נשלחה בהצלחה למכונאי.')
                
                repair_job.save()
                return redirect('repair_status', repair_id=repair_id)
            else:
                messages.error(request, 'נא לכתוב תגובה למכונאי.')
    
    updates = repair_job.updates.filter(is_visible_to_customer=True)
    
    # בדיקה אם יש פעולות תקועות או תיקון תקוע (למנהלים)
    has_blocked_items = repair_job.repair_items.filter(is_approved_by_customer=True, status='blocked').exists()
    show_manager_response = is_manager(request.user) and (repair_job.is_stuck or has_blocked_items)
    
    return render(request, 'workshop/repair_status.html', {
        'repair_job': repair_job,
        'updates': updates,
        'show_manager_response': show_manager_response,
    })

def send_customer_notification(repair_job, message_type, extra_message="", user=None):
    """שליחת התראה ללקוח על עדכון בתיקון"""
    customer = repair_job.bike.customer
    
    # אם יש למשתמש אימייל - נשלח אימייל
    if customer.email:
        subject_map = {
            'diagnosis_ready': f'אבחון מוכן עבור {repair_job.bike} - מוסך האופניים',
            'repair_completed': f'התיקון נשלם עבור {repair_job.bike} - מוסך האופניים',
            'general_update': f'עדכון בתיקון {repair_job.bike} - מוסך האופניים',
        }
        
        message_map = {
            'diagnosis_ready': f'''שלום {customer.name},

האבחון עבור האופניים {repair_job.bike} מוכן לצפייה ואישור.

פרטי האבחון:
{repair_job.diagnosis}

אנא היכנס למערכת כדי לאשר את הפעולות הנדרשות:
http://localhost:8000/repair/{repair_job.id}/approve/

{extra_message}

בברכה,
מוסך האופניים
''',
            'repair_completed': f'''שלום {customer.name},

התיקון עבור האופניים {repair_job.bike} הושלם בהצלחה!

ניתן לאסוף את האופניים במוסך.

בברכה,
מוסך האופניים
''',
            'general_update': f'''שלום {customer.name},

יש עדכון חדש בתיקון האופניים {repair_job.bike}.

{extra_message}

בברכה,
מוסך האופניים
'''
        }
        
        try:
            send_mail(
                subject=subject_map.get(message_type, 'עדכון ממוסך האופניים'),
                message=message_map.get(message_type, f'יש עדכון חדש: {extra_message}'),
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'garage@example.com'),
                recipient_list=[customer.email],
                fail_silently=True,  # לא נעצור את התהליך אם האימייל נכשל
            )
        except Exception as e:
            # נלוג את השגיאה אבל לא נעצור את התהליך
            # Failed to send email
            pass
    
    # תמיד נוסיף עדכון במערכת (רק אם יש user)
    if user:
        RepairUpdate.objects.create(
            repair_job=repair_job,
            user=user,
            message=f"נשלחה התראה ללקוח: {extra_message[:100]}..." if extra_message else f"התראה נשלחה - {message_type}",
            is_visible_to_customer=False  # זה עדכון פנימי
        )

@login_required
@manager_required
def customer_with_bike_form(request, customer_id=None, bike_id=None):
    """טופס משולב למנהל ליצירת/עריכת לקוח ואופניים"""
    
    customer_instance = None
    bike_instance = None
    
    if customer_id:
        customer_instance = get_object_or_404(Customer, id=customer_id)
        if bike_id:
            bike_instance = get_object_or_404(Bike, id=bike_id, customer=customer_instance)
        else:
            # אם יש לקוח אבל לא צוין bike_id, נלקח האופניים הראשונות
            bike_instance = customer_instance.bikes.first()
    
    if request.method == 'POST':
        form = CustomerWithBikeForm(
            request.POST,
            customer_instance=customer_instance,
            bike_instance=bike_instance
        )
        if form.is_valid():
            customer, bike = form.save()
            action = "עודכן" if customer_instance else "נוצר"
            messages.success(request, f'לקוח ואופניים {action}ו בהצלחה!')
            return redirect('customer_list')
    else:
        form = CustomerWithBikeForm(
            customer_instance=customer_instance,
            bike_instance=bike_instance
        )
    
    context = {
        'form': form,
        'customer_instance': customer_instance,
        'bike_instance': bike_instance,
        'title': 'עריכת לקוח ואופניים' if customer_instance else 'לקוח ואופניים חדשים'
    }
    return render(request, 'workshop/customer_with_bike_form.html', context)

@customer_required
@login_required
def customer_add_bike(request):
    """טופס ללקוח להוספת אופניים נוספות"""
    
    try:
        customer = Customer.objects.get(user=request.user)
    except Customer.DoesNotExist:
        messages.error(request, "לא נמצא פרופיל לקוח")
        return redirect('home')
    
    if request.method == 'POST':
        form = CustomerAddBikeForm(request.POST, customer=customer)
        if form.is_valid():
            bike = form.save()
            messages.success(request, f'האופניים {bike.brand} {bike.model} נוספו בהצלחה!')
            return redirect('home')
    else:
        form = CustomerAddBikeForm(customer=customer)
    
    context = {
        'form': form,
        'customer': customer,
        'title': 'הוספת אופניים נוספות'
    }
    return render(request, 'workshop/customer_add_bike.html', context)

@login_required
def customer_bikes_list(request):
    """רשימת האופניים של הלקוח"""
    
    try:
        customer = Customer.objects.get(user=request.user)
    except Customer.DoesNotExist:
        messages.error(request, "לא נמצא פרופיל לקוח")
        return redirect('home')
    
    bikes = customer.bikes.all()
    
    context = {
        'customer': customer,
        'bikes': bikes,
        'title': 'האופניים שלי'
    }
    return render(request, 'workshop/customer_bikes_list.html', context)


@login_required
@mechanic_required
def update_repair_status(request):
    """עדכון סטטוס תקוע של תיקון על ידי מכונאי"""
    if request.method == 'POST':
        try:
            repair_id = request.POST.get('repair_id')
            status = request.POST.get('status')
            reason = request.POST.get('reason', '')
            
            repair_job = get_object_or_404(RepairJob, id=repair_id, assigned_mechanic=request.user)
            
            if status == 'stuck':
                repair_job.is_stuck = True
                repair_job.stuck_reason = reason
                repair_job.stuck_at = timezone.now()
                repair_job.stuck_resolved = False
                repair_job.manager_response = ''
                
                # יצירת עדכון למעקב
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message=f"מכונאי דיווח על התקיעות: {reason}",
                    is_visible_to_customer=False
                )
                
                messages.success(request, 'התיקון סומן כתקוע. המנהל יקבל התראה.')
                
            elif status == 'working':
                repair_job.is_stuck = False
                repair_job.stuck_reason = ''
                repair_job.stuck_at = None
                repair_job.stuck_resolved = True
                
                # יצירת עדכון למעקב
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message="מכונאי חזר לעבודה רגילה",
                    is_visible_to_customer=False
                )
            
            repair_job.save()
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
@manager_required
def manager_response_stuck(request):
    """תגובת מנהל לתיקון תקוע"""
    if request.method == 'POST':
        try:
            repair_id = request.POST.get('repair_id')
            response = request.POST.get('response')
            mark_resolved = request.POST.get('mark_resolved') == 'true'
            
            # חיפוש התיקון - יכול להיות תקוע בגלל is_stuck או בגלל פעולות תקועות
            repair_job = get_object_or_404(RepairJob, id=repair_id)
            
            # בדיקה שהתיקון אכן תקוע
            has_blocked_items = repair_job.repair_items.filter(is_approved_by_customer=True, status='blocked').exists()
            is_repair_stuck = getattr(repair_job, 'is_stuck', False)
            
            if not (is_repair_stuck or has_blocked_items):
                return JsonResponse({'success': False, 'error': 'התיקון לא תקוע'})
            
            # עדכון התגובה
            repair_job.manager_response = response
            
            if mark_resolved:
                # סימון כנפתר
                if is_repair_stuck:
                    repair_job.stuck_resolved = True
                    repair_job.is_stuck = False
                
                # פתיחת פעולות תקועות (החזרה לסטטוס ממתין)
                blocked_items = repair_job.repair_items.filter(is_approved_by_customer=True, status='blocked')
                for item in blocked_items:
                    item.status = 'pending'
                    item.notes = item.notes + f"\n[מנהל פתר: {response}]" if item.notes else f"[מנהל פתר: {response}]"
                    item.save()
                
                # יצירת עדכון למעקב
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message=f"מנהל פתר את התקיעות: {response}",
                    is_visible_to_customer=False
                )
                
                message = 'התקיעות נפתרה בהצלחה'
            else:
                # יצירת עדכון למעקב
                RepairUpdate.objects.create(
                    repair_job=repair_job,
                    user=request.user,
                    message=f"מנהל השיב על התקיעות: {response}",
                    is_visible_to_customer=False
                )
                
                message = 'התגובה נשלחה בהצלחה'
            
            repair_job.save()
            
            return JsonResponse({'success': True, 'message': message})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
@login_required
def manager_quality_check(request, repair_id):
    """דף בדיקת איכות למנהל"""
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != 'manager':
        return redirect('login')
    
    # בדיקה אם השדות החדשים קיימים
    if not has_quality_fields():
        messages.error(request, 'תכונת בדיקת איכות עדיין לא זמינה במערכת')
        return redirect('manager_dashboard')
    
    repair_job = get_object_or_404(RepairJob, id=repair_id)
    
    # וודא שהתיקון מוכן לבדיקת איכות
    if repair_job.status != 'awaiting_quality_check':
        messages.error(request, 'תיקון זה אינו מוכן לבדיקת איכות')
        return redirect('manager_dashboard')
    
    return render(request, 'workshop/manager_quality_check.html', {
        'repair_job': repair_job,
        'approved_items': repair_job.approved_items,
        'completed_items': repair_job.completed_items,
    })


@login_required
def manager_quality_approve(request, repair_id):
    """אישור או דחיית בדיקת איכות על ידי מנהל"""
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != 'manager':
        return redirect('login')
    
    # בדיקה אם השדות החדשים קיימים
    if not has_quality_fields():
        messages.error(request, 'תכונת בדיקת איכות עדיין לא זמינה במערכת')
        return redirect('manager_dashboard')
    
    if request.method == 'POST':
        repair_job = get_object_or_404(RepairJob, id=repair_id)
        
        action = request.POST.get('action')
        quality_notes = request.POST.get('quality_notes', '')
        
        if action == 'approve':
            # אישור בדיקת איכות
            repair_job.status = 'quality_approved'
            repair_job.quality_checked_by = request.user
            repair_job.quality_check_date = timezone.now()
            repair_job.quality_notes = quality_notes
            repair_job.ready_for_pickup_date = timezone.now()
            repair_job.available_for_pickup = True
            repair_job.save()
            
            # Send push notification for ready for pickup
            from .push_service import push_service
            try:
                push_service.send_ready_for_pickup_notification(
                    repair_job.bike.customer, 
                    repair_job
                )
            except Exception as e:
                logger.error(f"Failed to send ready for pickup notification: {e}")
            
            # Send real-time notification to customer that repair is ready for pickup
            if repair_job.bike.customer.user:
                realtime_service.send_to_group(
                    f"customer_{repair_job.bike.customer.user.id}",
                    "repair_ready_for_pickup",
                    {
                        'repair_id': repair_job.id,
                        'bike_info': f"{repair_job.bike.brand} {repair_job.bike.model}",
                        'message': f"תיקון מוכן לאיסוף: {repair_job.bike.brand} {repair_job.bike.model}"
                    }
                )
            
            messages.success(request, f'תיקון #{repair_job.id} אושר ומוכן לאיסוף!')
            
        elif action == 'reject':
            # דחיית בדיקת איכות - החזרה למכונאי לתיקון
            repair_job.status = 'in_progress'
            repair_job.quality_notes = quality_notes
            repair_job.is_stuck = True
            repair_job.stuck_reason = f"דחיית בדיקת איכות: {quality_notes}"
            repair_job.stuck_at = timezone.now()
            repair_job.save()
            
            # Send real-time notification to mechanic about rejection
            if repair_job.assigned_mechanic:
                realtime_service.send_to_group(
                    f"mechanic_{repair_job.assigned_mechanic.id}",
                    "quality_check_rejected",
                    {
                        'repair_id': repair_job.id,
                        'bike_info': f"{repair_job.bike.brand} {repair_job.bike.model}",
                        'reason': quality_notes,
                        'message': f"בדיקת איכות נדחתה: {repair_job.bike.brand} {repair_job.bike.model}"
                    }
                )
            
            messages.warning(request, f'תיקון #{repair_job.id} נדחה בבדיקת האיכות והוחזר למכונאי')
        
        return redirect('manager_dashboard')
    
    return redirect('manager_quality_check', repair_id=repair_id)


@login_required
def manager_mark_delivered(request, repair_id):
    """סימון תיקון כנמסר ללקוח"""
    if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != 'manager':
        return JsonResponse({'success': False, 'error': 'אין הרשאה'})
    
    if request.method == 'POST':
        try:
            repair_job = get_object_or_404(RepairJob, id=repair_id)
            
            if repair_job.status != 'quality_approved':
                return JsonResponse({'success': False, 'error': 'תיקון זה אינו מוכן למסירה'})
            
            repair_job.status = 'delivered'
            repair_job.save()
            
            # יצירת עדכון למעקב
            RepairUpdate.objects.create(
                repair_job=repair_job,
                user=request.user,
                message="התיקון נמסר ללקוח בהצלחה.",
                is_visible_to_customer=True
            )
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
@user_passes_test(is_manager)
def manager_notify_customer(request, repair_id):
    """שליחת הודעה ללקוח שהתיקון מוכן לאיסוף"""
    if not has_quality_fields():
        return JsonResponse({'success': False, 'error': 'תכונת בדיקת איכות לא זמינה'})
    
    repair_job = get_object_or_404(RepairJob, id=repair_id)
    
    if repair_job.status != 'quality_approved':
        return JsonResponse({'success': False, 'error': 'התיקון לא עבר בדיקת איכות'})
    
    if request.method == 'POST':
        try:
            # עדכון שהלקוח הודע
            repair_job.customer_notified = True
            repair_job.save()
            
            # שליחת התראה ללקוח באמצעות מערכת ההתראות החדשה
            NotificationService.notify_ready_for_pickup(repair_job)
            
            # יצירת עדכון במערכת
            RepairUpdate.objects.create(
                repair_job=repair_job,
                user=request.user,
                message=f"הלקוח הודע שהתיקון מוכן לאיסוף. ניתן לבוא לאסוף את האופניים.",
                is_visible_to_customer=True
            )
            
            return JsonResponse({'success': True, 'message': 'הלקוח הודע בהצלחה'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'שיטה לא נתמכת'})


@login_required
@user_passes_test(lambda u: is_manager(u) or is_mechanic(u))
def print_bike_label(request, repair_id):
    """הדפסת מדבקה לאופניים עם פרטי התיקון"""
    repair_job = get_object_or_404(RepairJob, id=repair_id)
    
    return render(request, 'workshop/print_bike_label.html', {
        'repair_job': repair_job,
    })


@login_required
@user_passes_test(lambda u: is_manager(u) or is_mechanic(u))
def print_labels_menu(request):
    """תפריט הדפסת מדבקות - רשימת תיקונים פעילים"""
    
    # תיקונים פעילים לפי קטגוריות
    active_repairs = RepairJob.objects.filter(
        status__in=['reported', 'diagnosed', 'approved', 'in_progress', 'awaiting_quality_check']
    ).select_related('bike', 'bike__customer', 'assigned_mechanic').order_by('-created_at')
    
    # קיבוץ לפי סטטוס
    repairs_by_status = {}
    for repair in active_repairs:
        status = repair.get_status_display()
        if status not in repairs_by_status:
            repairs_by_status[status] = []
        repairs_by_status[status].append(repair)
    
    return render(request, 'workshop/print_labels_menu.html', {
        'repairs_by_status': repairs_by_status,
        'total_repairs': active_repairs.count(),
    })


@login_required
@manager_required
def backup_customer_data(request):
    """גיבוי נתוני לקוחות - יצוא לקובץ Excel"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    import io
    
    # יצירת workbook חדש
    wb = openpyxl.Workbook()
    
    # גיליון לקוחות
    ws_customers = wb.active
    ws_customers.title = "לקוחות"
    
    # כותרות עבור גיליון לקוחות
    customer_headers = [
        'מספר לקוח', 'שם', 'טלפון', 'אימייל', 
        'תאריך הרשמה', 'משתמש מקושר', 'הערות'
    ]
    
    # הוספת כותרות
    for col, header in enumerate(customer_headers, 1):
        cell = ws_customers.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # הוספת נתוני לקוחות
    customers = Customer.objects.select_related('user').all().order_by('name')
    for row, customer in enumerate(customers, 2):
        ws_customers.cell(row=row, column=1, value=customer.id)
        ws_customers.cell(row=row, column=2, value=customer.name)
        ws_customers.cell(row=row, column=3, value=customer.phone or '')
        ws_customers.cell(row=row, column=4, value=customer.email or '')
        ws_customers.cell(row=row, column=5, value=customer.created_at.strftime('%d/%m/%Y %H:%M') if hasattr(customer, 'created_at') else '')
        ws_customers.cell(row=row, column=6, value=customer.user.username if customer.user else 'לא מקושר')
        ws_customers.cell(row=row, column=7, value=getattr(customer, 'notes', '') or '')
    
    # התאמת רוחב עמודות
    for col in range(1, len(customer_headers) + 1):
        ws_customers.column_dimensions[get_column_letter(col)].width = 15
    
    # גיליון אופניים
    ws_bikes = wb.create_sheet(title="אופניים")
    bike_headers = [
        'מספר אופניים', 'מספר לקוח', 'שם לקוח', 'מותג', 'דגם', 
        'צבע', 'הערות'
    ]
    
    for col, header in enumerate(bike_headers, 1):
        cell = ws_bikes.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    bikes = Bike.objects.select_related('customer').all().order_by('customer__name', 'brand')
    for row, bike in enumerate(bikes, 2):
        ws_bikes.cell(row=row, column=1, value=bike.id)
        ws_bikes.cell(row=row, column=2, value=bike.customer.id)
        ws_bikes.cell(row=row, column=3, value=bike.customer.name)
        ws_bikes.cell(row=row, column=4, value=bike.brand or '')
        ws_bikes.cell(row=row, column=5, value=bike.model or '')
        ws_bikes.cell(row=row, column=6, value=bike.color or '')
        ws_bikes.cell(row=row, column=7, value=getattr(bike, 'notes', '') or '')
    
    for col in range(1, len(bike_headers) + 1):
        ws_bikes.column_dimensions[get_column_letter(col)].width = 15
    
    # גיליון תיקונים
    ws_repairs = wb.create_sheet(title="תיקונים")
    repair_headers = [
        'מספר תיקון', 'מספר אופניים', 'שם לקוח', 'אופניים', 'בעיה',
        'אבחון', 'סטטוס', 'מכונאי', 'תאריך קבלה', 'תאריך אבחון',
        'מחיר כולל', 'הערות'
    ]
    
    for col, header in enumerate(repair_headers, 1):
        cell = ws_repairs.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D83B01", end_color="D83B01", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    repairs = RepairJob.objects.select_related(
        'bike', 'bike__customer', 'assigned_mechanic'
    ).prefetch_related('repair_items').all().order_by('-created_at')
    
    for row, repair in enumerate(repairs, 2):
        ws_repairs.cell(row=row, column=1, value=repair.id)
        ws_repairs.cell(row=row, column=2, value=repair.bike.id)
        ws_repairs.cell(row=row, column=3, value=repair.bike.customer.name)
        ws_repairs.cell(row=row, column=4, value=f"{repair.bike.brand} {repair.bike.model or ''}".strip())
        ws_repairs.cell(row=row, column=5, value=repair.problem_description or '')
        ws_repairs.cell(row=row, column=6, value=repair.diagnosis or '')
        ws_repairs.cell(row=row, column=7, value=repair.get_status_display())
        ws_repairs.cell(row=row, column=8, value=repair.assigned_mechanic.get_full_name() if repair.assigned_mechanic else '')
        ws_repairs.cell(row=row, column=9, value=repair.created_at.strftime('%d/%m/%Y %H:%M'))
        ws_repairs.cell(row=row, column=10, value=repair.diagnosed_at.strftime('%d/%m/%Y %H:%M') if repair.diagnosed_at else '')
        
        # חישוב מחיר כולל
        total_price = sum(item.price for item in repair.repair_items.all())
        ws_repairs.cell(row=row, column=11, value=total_price)
        
        # הערות מפעולות התיקון
        repair_notes = '; '.join([item.notes for item in repair.repair_items.all() if item.notes])
        ws_repairs.cell(row=row, column=12, value=repair_notes or '')
    
    for col in range(1, len(repair_headers) + 1):
        ws_repairs.column_dimensions[get_column_letter(col)].width = 15
    
    # גיליון סטטיסטיקות
    ws_stats = wb.create_sheet(title="סטטיסטיקות")
    
    # כותרת
    ws_stats.cell(row=1, column=1, value="דוח סטטיסטיקות - גיבוי נתונים")
    ws_stats.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws_stats.cell(row=2, column=1, value=f"נוצר ב: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # סטטיסטיקות
    stats = [
        ("סך הכל לקוחות:", customers.count()),
        ("לקוחות עם משתמש מקושר:", customers.filter(user__isnull=False).count()),
        ("לקוחות ללא משתמש:", customers.filter(user__isnull=True).count()),
        ("סך הכל אופניים:", bikes.count()),
        ("סך הכל תיקונים:", repairs.count()),
        ("תיקונים פעילים:", repairs.filter(status__in=['reported', 'diagnosed', 'approved', 'in_progress']).count()),
        ("תיקונים שהושלמו:", repairs.filter(status__in=['completed', 'delivered']).count()),
    ]
    
    for i, (label, value) in enumerate(stats, 4):
        ws_stats.cell(row=i, column=1, value=label)
        ws_stats.cell(row=i, column=2, value=value)
        ws_stats.cell(row=i, column=1).font = Font(bold=True)
    
    # התאמת רוחב עמודות
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 15
    
    # שמירה לזיכרון
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # יצירת תגובה להורדה
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"backup_customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@manager_required  
def backup_menu(request):
    """תפריט גיבויים - אפשרויות גיבוי שונות"""
    
    # סטטיסטיקות לתצוגה
    customers_count = Customer.objects.count()
    bikes_count = Bike.objects.count()
    repairs_count = RepairJob.objects.count()
    active_repairs = RepairJob.objects.filter(
        status__in=['reported', 'diagnosed', 'approved', 'in_progress']
    ).count()
    
    context = {
        'customers_count': customers_count,
        'bikes_count': bikes_count,
        'repairs_count': repairs_count,
        'active_repairs': active_repairs,
    }
    
    return render(request, 'workshop/backup_menu.html', context)


@login_required
@manager_required
def backup_customers_csv(request):
    """גיבוי לקוחות פשוט בפורמט CSV"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"customers_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # הוספת BOM עבור UTF-8 כדי שExcel יזהה עברית
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # כותרות
    writer.writerow([
        'מספר לקוח', 'שם', 'טלפון', 'אימייל', 
        'משתמש מקושר', 'מספר אופניים', 'סה"כ תיקונים'
    ])
    
    # נתוני לקוחות
    customers = Customer.objects.select_related('user').prefetch_related('bikes', 'bikes__repairjob_set').all()
    
    for customer in customers:
        bikes_count = customer.bikes.count()
        repairs_count = sum(bike.repairjob_set.count() for bike in customer.bikes.all())
        
        writer.writerow([
            customer.id,
            customer.name,
            customer.phone or '',
            customer.email or '',
            customer.user.username if customer.user else 'לא מקושר',
            bikes_count,
            repairs_count
        ])
    
    return response


# API endpoints for repair form customer search
@login_required
@user_passes_test(lambda u: is_manager(u) or is_mechanic(u))
def search_customers_api(request):
    """API endpoint for searching customers"""
    try:
        query = request.GET.get('q', '').strip()
        if len(query) < 2:
            return JsonResponse({'customers': []})
        
        # Search customers by name, phone, email, or associated username
        customers = Customer.objects.filter(
            Q(name__icontains=query) |
            Q(phone__icontains=query) |
            Q(email__icontains=query) |
            Q(user__username__icontains=query)  # Added username search
        ).select_related('user').prefetch_related('bikes')[:10]  # Limit to 10 results
        
        
        customers_data = []
        for customer in customers:
            customer_data = {
                'id': customer.id,
                'name': customer.name,
                'phone': customer.phone or '',
                'email': customer.email or '',
                'bikes_count': customer.bikes.count(),
                'username': customer.user.username if customer.user else ''
            }
            customers_data.append(customer_data)
        
        return JsonResponse({'customers': customers_data})
    
    except Exception as e:
        # Error in search_customers_api
        return JsonResponse({'error': f'Search error: {str(e)}'}, status=500)


@login_required
@user_passes_test(lambda u: is_manager(u) or is_mechanic(u))
def customer_bikes_api(request, customer_id):
    """API endpoint for getting customer's bikes"""
    try:
        customer = Customer.objects.get(id=customer_id)
        bikes = customer.bikes.prefetch_related('repairjob_set').all()
        
        bikes_data = []
        for bike in bikes:
            bikes_data.append({
                'id': bike.id,
                'brand': bike.brand,
                'model': bike.model,
                'color': bike.color,
                'repairs_count': bike.repairjob_set.count()
            })
        
        return JsonResponse({'bikes': bikes_data})
    
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)


def csrf_failure(request, reason=""):
    """Custom CSRF failure view to handle CSRF errors gracefully"""
    messages.error(request, 'שגיאת אבטחה. אנא נסה שוב.')
    return redirect('login')



